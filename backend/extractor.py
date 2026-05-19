import io
from pypdf import PdfReader
from docx import Document
from openpyxl import load_workbook

SUPPORTED_TYPES = {".txt", ".pdf", ".docx", ".xlsx"}

def extract_text(file_bytes: bytes, file_extension: str) -> str:
    """
    Extract plain text from a file given its raw bytes and extension.
    Returns empty string if extraction fails or type is unsupported.
    """
    ext = file_extension.lower()

    try:
        if ext == ".txt":
            return _extract_txt(file_bytes)
        elif ext == ".pdf":
            return _extract_pdf(file_bytes)
        elif ext == ".docx":
            return _extract_docx(file_bytes)
        elif ext == ".xlsx":
            return _extract_xlsx(file_bytes)
        else:
            return ""
    except Exception as e:
        print(f"Extraction failed for {file_extension}: {e}")
        return ""

def _extract_txt(file_bytes: bytes) -> str:
    try:
        return file_bytes.decode("utf-8")
    except UnicodeDecodeError:
        return file_bytes.decode("latin-1", errors="ignore")

def _extract_pdf(file_bytes: bytes) -> str:
    reader = PdfReader(io.BytesIO(file_bytes))
    pages = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            pages.append(text)
    return "\n".join(pages)

def _extract_docx(file_bytes: bytes) -> str:
    doc = Document(io.BytesIO(file_bytes))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n".join(paragraphs)

def _extract_xlsx(file_bytes: bytes) -> str:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    lines = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            row_text = " ".join(str(cell) for cell in row if cell is not None)
            if row_text.strip():
                lines.append(row_text)
    return "\n".join(lines)